import io
import openpyxl

from decimal import Decimal
from datetime import date
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.db.models import Max
from reportlab.platypus import Image
from reportlab.lib.utils import ImageReader
import os

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
)
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from apps.boq.models import BOQ, BOQItem
from apps.projects.models import Project
from apps.configurations.models import (
    LightingConfiguration,
    ConfigurationAccessory,
    ConfigurationDriver
)

class BOQPDFBuilder:
    def __init__(self, boq, is_draft=False):
        self.boq = boq
        self.is_draft = is_draft
        self.buffer = io.BytesIO()

        # Landscape A4
        self.pagesize = landscape(A4)
        self.width, self.height = self.pagesize

        self.MARGIN_X = 12 * mm
        self.MARGIN_Y = 12 * mm

        # Register Unicode font (for ₹ symbol)
        font_path = os.path.join("fonts", "DejaVuSans.ttf")
        pdfmetrics.registerFont(TTFont("DejaVu", font_path))

        self.styles = getSampleStyleSheet()

        self.style_normal = ParagraphStyle(
            'NormalSmall',
            parent=self.styles['Normal'],
            fontName='DejaVu',
            fontSize=9,
            leading=12
        )

        self.style_bold = ParagraphStyle(
            'BoldSmall',
            parent=self.styles['Normal'],
            fontName='DejaVu',
            fontSize=9,
            leading=12,
        )

    # --------------------------------------------------
    # Header + Footer
    # --------------------------------------------------
    def _header_footer(self, canvas, doc):
        canvas.saveState()

        canvas.setFont('DejaVu', 14)
        canvas.drawString(self.MARGIN_X, self.height - 30, "TVUM TECH")

        canvas.setFont('DejaVu', 10)
        canvas.drawString(self.MARGIN_X, self.height - 45, "Lighting ERP – Bill of Quantities")

        canvas.setFont('DejaVu', 9)
        canvas.drawString(self.MARGIN_X, self.height - 65, f"Project: {self.boq.project.name}")
        canvas.drawRightString(
            self.width - self.MARGIN_X,
            self.height - 65,
            f"BOQ Version: {self.boq.version}"
        )

        canvas.drawString(self.MARGIN_X, self.height - 80, f"Status: {self.boq.status}")
        canvas.drawRightString(
            self.width - self.MARGIN_X,
            self.height - 80,
            f"Date: {date.today().strftime('%d-%m-%Y')}"
        )

        canvas.setStrokeColor(colors.black)
        canvas.line(
            self.MARGIN_X,
            self.height - 90,
            self.width - self.MARGIN_X,
            self.height - 90
        )

        canvas.setFont('DejaVu', 8)
        canvas.drawCentredString(
            self.width / 2,
            15,
            "System Generated BOQ | TVUM Lighting ERP"
        )
        canvas.drawRightString(
            self.width - self.MARGIN_X,
            15,
            f"Page {doc.page}"
        )

        if self.is_draft:
            self._draw_watermark(canvas)

        canvas.restoreState()

    # --------------------------------------------------
    # Watermark
    # --------------------------------------------------
    def _draw_watermark(self, canvas):
        canvas.saveState()
        canvas.setFont('DejaVu', 50)
        canvas.setFillColor(colors.lightgrey, alpha=0.15)
        canvas.translate(self.width / 2, self.height / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, "DRAFT - FOR INTERNAL USE ONLY")
        canvas.restoreState()

    # --------------------------------------------------
    # Currency
    # --------------------------------------------------
    def _format_currency(self, value):
        return f"₹ {Decimal(value):,.2f}"

    # --------------------------------------------------
    # Main Builder
    # --------------------------------------------------
    def build(self):
        doc = SimpleDocTemplate(
            self.buffer,
            pagesize=self.pagesize,
            leftMargin=20 * mm,
            rightMargin=20 * mm,
            topMargin=35 * mm,
            bottomMargin=25 * mm
        )

        elements = []
        grand_total = Decimal(0)

        areas = (
            self.boq.items
            .select_related("area")
            .values_list("area__id", "area__name")
            .distinct()
        )

        for area_id, area_name in areas:
            elements.append(Paragraph(f"<b>Area: {area_name}</b>", self.styles['Heading4']))
            elements.append(Spacer(1, 5))

            table_data = [[
                "Image",
                "Type",
                "Code",
                "Description",
                "Qty",
                "Unit",
                "Rate",
                "GST",
                "Total"
            ]]

            area_total = Decimal(0)
            items = self.boq.items.filter(area_id=area_id)

            for item in items:

                qty = Decimal(item.quantity or 0)
                line_total = Decimal(item.final_price or 0)

                # derive base for display
                if qty > 0:
                    total_unit = line_total / qty
                else:
                    total_unit = Decimal("0")

                base_price = total_unit / Decimal("1.18")
                gst = total_unit - base_price

                area_total += line_total
                grand_total += line_total

                img = ""
                item_code = "-"
                desc = "-"

                # PRODUCT
                if item.item_type == "PRODUCT" and item.product:
                    product = item.product
                    item_code = product.order_code or "-"
                    desc = (
                        f"{product.make} | "
                        f"{product.wattage}W | "
                        f"{product.cct_kelvin}K | "
                        f"{product.beam_angle_degree}°"
                    )

                    if product.visual_image:
                        try:
                            img = Image(
                                product.visual_image.path,
                                width=28,
                                height=28
                            )
                        except:
                            img = ""

                    table_data.append([
                        img,
                        "Product",
                        item_code,
                        Paragraph(desc, self.style_normal),
                        str(qty),
                        "Nos",
                        self._format_currency(base_price),
                        self._format_currency(gst),
                        self._format_currency(line_total)
                    ])

                # DRIVER
                elif item.item_type == "DRIVER" and item.driver:
                    driver = item.driver
                    desc = f"{driver.driver_make} {driver.driver_code}"

                    table_data.append([
                        "",
                        Paragraph("→ Driver", self.style_normal),
                        "-",
                        Paragraph(desc, self.style_normal),
                        str(qty),
                        "Nos",
                        self._format_currency(base_price),
                        self._format_currency(gst),
                        self._format_currency(line_total)
                    ])

                # ACCESSORY
                elif item.item_type == "ACCESSORY" and item.accessory:
                    acc = item.accessory
                    desc = acc.accessory_name

                    table_data.append([
                        "",
                        Paragraph("→ Accessory", self.style_normal),
                        "-",
                        Paragraph(desc, self.style_normal),
                        str(qty),
                        "Nos",
                        self._format_currency(base_price),
                        self._format_currency(gst),
                        self._format_currency(line_total)
                    ])

            # Area subtotal
            table_data.append([
                "", "", "", "Area Subtotal",
                "", "", "",
                "",
                self._format_currency(area_total)
            ])

            usable_width = doc.width
            table = Table(
                table_data,
                colWidths=[
                    usable_width * 0.05,  # image
                    usable_width * 0.08,  # type
                    usable_width * 0.12,  # code
                    usable_width * 0.30,  # description
                    usable_width * 0.07,  # qty
                    usable_width * 0.07,  # unit
                    usable_width * 0.10,  # rate
                    usable_width * 0.10,  # gst
                    usable_width * 0.11,  # total
                ],
                repeatRows=1,
                hAlign="CENTER"
            )

            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
                ('FONTNAME', (0, 0), (-1, 0), 'DejaVu'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),

                ('FONTNAME', (0, 1), (-1, -1), 'DejaVu'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),

                ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

                # internal padding
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),

                ('ROWBACKGROUNDS', (0, 1), (-1, -2),
                [colors.white, colors.HexColor("#F7F7F7")]),

                ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),

                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#EFEFEF")),
                ('FONTNAME', (0, -1), (-1, -1), 'DejaVu'),
            ]))

            elements.append(table)
            elements.append(Spacer(1, 15))

        # Grand total
        elements.append(Paragraph(
            f"Grand Total: {self._format_currency(grand_total)}",
            ParagraphStyle(
                'Total',
                parent=self.styles['Normal'],
                alignment=TA_RIGHT,
                fontName='DejaVu',
                fontSize=13,
            )
        ))

        doc.build(
            elements,
            onFirstPage=self._header_footer,
            onLaterPages=self._header_footer
        )

        self.buffer.seek(0)
        response = HttpResponse(self.buffer, content_type="application/pdf")
        filename = f"{self.boq.project.name}_V{self.boq.version}_{self.boq.status}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

@transaction.atomic
def generate_boq(project, user):
    """
    ERP-grade BOQ generation.

    Rules:
    - BOQ generated only from active configuration
    - One BOQ per configuration version
    - Supports project-level and area-wise configs
    - Append-only versioning
    """

    # -----------------------------
    # 1. LOAD PROJECT SAFELY
    # -----------------------------
    project_id = None

    if hasattr(project, "id"):
        project_id = project.id
    elif isinstance(project, dict):
        project_id = project.get("id")
    else:
        project_id = project

    if not project_id:
        raise ValidationError("Invalid project reference")

    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        raise ValidationError("Project not found")

    # -----------------------------
    # 2. LOAD ACTIVE CONFIGURATIONS
    # -----------------------------
    active_configs = LightingConfiguration.objects.filter(
        project=project,
        is_active=True
    ).select_related("area", "product")

    if not active_configs.exists():
        raise ValidationError(
            "No active configurations found. "
            "Please save configuration first."
        )

    # All configs must share same version
    config_version = active_configs.first().configuration_version

    # -----------------------------
    # 3. PREVENT DUPLICATE BOQ
    # -----------------------------
    latest_boq = BOQ.objects.filter(project=project).order_by("-version").first()

    if latest_boq and latest_boq.source_configuration_version == config_version:
        raise ValidationError(
            "BOQ already generated for the current configuration version."
        )

    # -----------------------------
    # 4. DETERMINE NEXT BOQ VERSION
    # -----------------------------
    latest_version = (
        BOQ.objects.filter(project=project)
        .aggregate(max_v=Max("version"))
        .get("max_v") or 0
    )

    next_boq_version = latest_version + 1

    # -----------------------------
    # 5. CREATE BOQ HEADER
    # -----------------------------
    boq = BOQ.objects.create(
    project=project,
    version=next_boq_version,
    created_by=user,
    status="DRAFT",
    source_configuration_version=config_version,
    )

    # -----------------------------
    # 6. CREATE BOQ ITEMS
    # -----------------------------
    seen_products = set()

    for config in active_configs:

        area = config.area
        product = config.product

        # Prevent duplicates
        area_key = area.id if area else "PROJECT"
        key = (area_key, product.prod_id)

        if key in seen_products:
            continue

        seen_products.add(key)

        # -------------------------
        # PRODUCT ITEM
        # -------------------------
        product_price = getattr(product, "base_price", 0)
        product_total = product_price * config.quantity

        product_item = BOQItem.objects.create(
            boq=boq,
            area=area,
            item_type="PRODUCT",
            product=product,
            quantity=config.quantity,
            unit_price=product_price,
            markup_pct=0,
            final_price=product_total,
        )

        # -------------------------
        # DRIVER ITEMS
        # -------------------------
        drivers = ConfigurationDriver.objects.filter(configuration=config)

        for drv in drivers:
            driver = drv.driver
            driver_price = getattr(driver, "base_price", 0)
            driver_total = driver_price * drv.quantity

            BOQItem.objects.create(
                boq=boq,
                area=area,
                item_type="DRIVER",
                driver=driver,
                quantity=drv.quantity,
                unit_price=driver_price,
                markup_pct=0,
                final_price=driver_total,
            )

        # -------------------------
        # ACCESSORY ITEMS
        # -------------------------
        accessories = ConfigurationAccessory.objects.filter(configuration=config)

        for acc in accessories:
            accessory = acc.accessory
            acc_price = getattr(accessory, "base_price", 0)
            acc_total = acc_price * acc.quantity

            BOQItem.objects.create(
                boq=boq,
                area=area,
                item_type="ACCESSORY",
                accessory=accessory,
                quantity=acc.quantity,
                unit_price=acc_price,
                markup_pct=0,
                final_price=acc_total,
            )

    # -----------------------------
    # 7. RETURN RESULT
    # -----------------------------
    return {
        "boq_id": boq.id,
        "version": boq.version,
        "status": boq.status,
        "source_configuration_version": config_version,
    }


def get_boq_summary(boq):
    if not boq:
        return None

    # 🔥 cumulative items
    items = BOQItem.objects.filter(
        boq__project=boq.project,
        boq__version__lte=boq.version
    ).values("item_type").annotate(
        total_qty=Sum("quantity"),
        total_value=Sum("final_price")
    )

    summary = {}
    for item in items:
        summary[item["item_type"]] = {
            "quantity": item["total_qty"],
            "amount": float(item["total_value"] or 0)
        }

    return {
        "project_id": boq.project.id,
        "boq_id": boq.id,
        "version": boq.version,
        "status": boq.status,
        "summary": summary,
        "created_at": boq.created_at,
        "source_configuration_version": boq.source_configuration_version
    }


def get_project_boq_summary(project):
    boq = BOQ.objects.filter(project=project).order_by("-version").first()
    return get_boq_summary(boq)

def approve_boq(boq):
    if boq.status != "DRAFT":
        raise ValidationError("Already finalized")
    boq.status = "FINAL"
    boq.locked_at = timezone.now()
    boq.save()
    return boq

def apply_margin_to_boq(boq, markup_pct):
    if boq.status != "DRAFT":
        raise ValidationError("Cannot modify FINAL BOQ")
    markup_pct = Decimal(markup_pct)
    for item in boq.items.all():
        item.markup_pct = markup_pct
        item.final_price = Decimal(item.unit_price) * Decimal(item.quantity) * (Decimal(1) + markup_pct / Decimal(100))
        item.save(update_fields=["markup_pct", "final_price"])
    return boq

class BOQExcelBuilder:
    def __init__(self, boq):
        self.boq = boq
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "BOQ"
        self.row = 1

    def money(self, value): return float(round(Decimal(value), 2))
    def bold(self): return openpyxl.styles.Font(bold=True)
    def center(self): return openpyxl.styles.Alignment(horizontal="center")
    def right(self): return openpyxl.styles.Alignment(horizontal="right")

    def write(self, col, value, bold=False, align=None, currency=False):
        cell = self.ws.cell(row=self.row, column=col, value=value)
        if bold: cell.font = self.bold()
        if align: cell.alignment = align
        if currency: cell.number_format = '₹#,##0.00'
        return cell

    def build(self):
        if self.boq.status != "FINAL": raise ValidationError("Excel export is allowed only for FINAL BOQ")
        self.write(1, "TVUM TECH", bold=True)
        self.row += 1
        self.write(1, "Lighting ERP – Bill of Quantities", bold=True)
        self.row += 2
        self.write(1, "Project:", bold=True); self.write(2, self.boq.project.name)
        self.write(5, "Version:", bold=True); self.write(6, self.boq.version)
        self.row += 1
        self.write(1, "Status:", bold=True); self.write(2, self.boq.status)
        self.write(5, "Date:", bold=True); self.write(6, date.today().strftime("%d-%m-%Y"))
        self.row += 2
        grand_total = Decimal(0)
        areas = self.boq.items.select_related("area").order_by("area__name").values_list("area__id", "area__name").distinct()
        for area_id, area_name in areas:
            self.write(1, f"Area: {area_name}", bold=True); self.row += 1
            headers = ["Type", "Item Code", "Description", "Qty", "Unit Price (₹)", "Margin (%)", "Line Total (₹)"]
            for col, h in enumerate(headers, start=1): self.write(col, h, bold=True, align=self.center())
            self.row += 1
            area_total = Decimal(0)
            for item in self.boq.items.filter(area_id=area_id):
                qty = Decimal(item.quantity)
                unit_price = Decimal(item.unit_price or 0)
                margin = Decimal(item.markup_pct or 0)
                line_total = Decimal(item.final_price or 0)
                area_total += line_total
                grand_total += line_total
                desc = "-"
                if item.item_type == "PRODUCT" and item.product: desc = item.product.make
                elif item.item_type == "DRIVER" and item.driver: desc = item.driver.driver_type
                elif item.item_type == "ACCESSORY" and item.accessory: desc = item.accessory.accessory_type
                self.write(1, item.item_type)
                self.write(2, item.product.order_code if item.product else "-")
                self.write(3, desc)
                self.write(4, int(qty), align=self.center())
                self.write(5, self.money(unit_price), align=self.right(), currency=True)
                self.write(6, float(margin), align=self.center())
                self.write(7, self.money(line_total), align=self.right(), currency=True)
                self.row += 1
            self.write(6, "Area Total", bold=True, align=self.right())
            self.write(7, self.money(area_total), bold=True, currency=True)
            self.row += 2
        self.write(6, "Grand Total", bold=True, align=self.right())
        self.write(7, self.money(grand_total), bold=True, currency=True)
        for col in range(1, 8): self.ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="BOQ_{self.boq.project.name}_V{self.boq.version}.xlsx"'
        self.wb.save(response)
        return response